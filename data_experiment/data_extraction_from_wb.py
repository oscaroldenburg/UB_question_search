import json
import openpyxl
from pathlib import Path
import colorsys
import re




EXCEL_FILE = "../api/Frågenyckel - Telekom 02-25 - test.xlsx"
OUTPUT_JSON = "questions.json"
ANSWER_TYPE_MAP = {
    # 🟩 singel – båda gröna varianter
    "FF92D050": "singel",
    "FF00FF00": "singel",   # om denna förekommer

    # 🟦 multi – blå toner
    "FF3333FF": "multi",    # ljusare blå
    "FF002060": "multi",    # mörkare blå

    # 🟪 numerisk
    "FF7030A0": "numerisk",

    # 🟨 loop
    "FFFFC000": "loop",

    # 🔴 skala
    "FFFF0000": "skala",
    "FF0000D4": "skala",

    # ⚫ 3D-skala
    "FF000000": "skala_3d",

    # 🟡 fritext
    "FFFFFF00": "öppen",

    # ⚙️ uteslutande (gissad grå)
    "FF7F7F7F": "uteslutande",

    "00000000": None,  # ingen färg
}

def hex_to_rgb(hex_color):
    """Convert hex color (e.g., 'FF92D050') to RGB tuple."""
    # Remove alpha channel if present (first 2 chars in format like FF92D050)
    if len(hex_color) == 8:
        hex_color = hex_color[2:]
    
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    return (r, g, b)

def color_distance(rgb1, rgb2):
    """Calculate Euclidean distance between two RGB colors."""
    return sum((a - b) ** 2 for a, b in zip(rgb1, rgb2)) ** 0.5

def classify_color(hex_color):
    """Classify a hex color into a question type category."""
    
    # Define reference colors for each category
    color_categories = {
        "singel": (146, 208, 80),          # Green (92D050)
        "multi": (51, 51, 255),            # Blue (3333FF)
        "skala": (255, 0, 0),              # Red (FF0000)
        "skala_3d": (0, 0, 0),             # Black (000000)
        "öppen": (255, 255, 0),            # Yellow (FFFF00)
        "uteslutande": (127, 127, 127),    # Gray (7F7F7F)
        "numerisk": (112, 48, 160),        # Purple (7030A0)
        "loop": (255, 192, 0),             # Orange/Gold (FFC000)
    }
    
    rgb = hex_to_rgb(hex_color)
    
    # Find the closest color category
    min_distance = float('inf')
    closest_category = None
    
    for category, ref_color in color_categories.items():
        distance = color_distance(rgb, ref_color)
        if distance < min_distance:
            min_distance = distance
            closest_category = category
    
    return closest_category
def get_fill_color(cell):
    """
    Försöker extrahera en RGB-färgkod från cellens fyllning.
    Returnerar kategori baserat på färg eller None om ingen färg hittas.
    """
    fill = cell.fill
    if fill is None:
        print("No fill found")
        return None

    # openpyxl använder fgColor / start_color
    color = getattr(fill, "fgColor", None)
    if color is not None and getattr(color, "type", None) == "rgb":
        hex_color = color.rgb
        if hex_color == "00000000":  # No fill
            print("No fill color detected")
            return None
        return classify_color(hex_color)

    start_color = getattr(fill, "start_color", None)
    if start_color is not None and getattr(start_color, "type", None) == "rgb":
        hex_color = start_color.rgb
        if hex_color == "00000000":  # No fill
            return None
        return classify_color(hex_color)

    return None
def get_fill_colorOLD(cell):
    """
    Försöker extrahera en RGB-färgkod från cellens fyllning.
    Returnerar t.ex. 'FFFF0000' eller None om ingen färg hittas.
    """
    """ answerTypeMap = {
        "singel":  """
    
    fill = cell.fill
    if fill is None:
        return None

    # openpyxl använder fgColor / start_color
    color = getattr(fill, "fgColor", None)
    if color is not None and getattr(color, "type", None) == "rgb":
        try:
            return ANSWER_TYPE_MAP[color.rgb]
        except KeyError:
            print("Unknown color code:", color.rgb)

    start_color = getattr(fill, "start_color", None)
    if start_color is not None and getattr(start_color, "type", None) == "rgb":
        try:
            return ANSWER_TYPE_MAP[start_color.rgb]
        except KeyError:
            print("Unknown start color code:", start_color.rgb)

    return None
    


def extract_questions_from_workbook(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    pattern = re.compile(r".*\^f\('([^']+)'\)\^.*")
    all_questions = []
    ws = wb.active
    var_map = {}
    question_index = 0

    for ws in wb.worksheets:
        category = ws.title
        max_row = ws.max_row
        max_col = ws.max_column


        for row in range(1, max_row + 1):
            
            q_cell = ws[f"F{row}"]
            value = q_cell.value

            if value is None:
                continue

            text = str(value).strip()
            is_bold = bool(q_cell.font and q_cell.font.bold)

            # Villkor för att en cell ska räknas som fråga
            if not is_bold:
                continue
            if not ("?" in text or "..." in text or "…" in text):
                continue

            question = text
            match = re.search(r"\^f\('([^']+)'\)\^", question)
            if match:
                refference_name = match.group(1)
                pritty_question = question.replace(match.group(1), "(variabel)").question.replace(match.group(0), "").question.replace(match.group(2), "").strip()







            # --- Hämta svarsalternativ ---
            answer_alternatives = []
            alt_row = row
            answer_type_color = get_fill_color(ws[f"D{row}"])

            if alt_row >= 1 and answer_type_color != "öppen":

                # Kolumn H = index 8, sedan åt höger tills tom cell
                while True:
                    alt_cell = ws.cell(row=alt_row, column=8)
                    alt_val = alt_cell.value

                    if alt_val is None or str(alt_val).strip() == "":
                        if alt_row == row:
                            alt_row += 1
                            continue
                        else:
                            break
                    answer_alternatives.append(str(alt_val).strip())
                    alt_row += 1

            # Free_Text: True om inga svarsalternativ hittades

            # Answer_type via färg i kolumn D på samma rad
            var_name = ws[f"E{row}"].value

            question_obj = {
                "question": question,
                "pritty_question": pritty_question if match else question,
                "answer_alternatives": answer_alternatives,
                "category": category,
                "answer_type": answer_type_color,  # här kan du senare mappa färg -> "singel"/"multi"
                "refference_name": refference_name if match else None,
                "variable_name": var_name,
            }

            all_questions.append(question_obj)
            if var_name:
                var_map[var_name] = question_index
            question_index += 1


    return all_questions, var_map


def main():
    xlsx_path = Path(EXCEL_FILE)
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"Hittar inte filen: {xlsx_path.resolve()}")

    data, var_map = extract_questions_from_workbook(str(xlsx_path))

    # Spara till JSON
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"✅ Klart! Extraherade {len(data)} frågor.")
    print(f"JSON sparad som: {OUTPUT_JSON}")

    # Spara var_map till JSON
    var_map_json = "var_map.json"
    with open(var_map_json, "w", encoding="utf-8") as f:
        json.dump(var_map, f, ensure_ascii=False, indent=2)

    print(f"✅ Klart! Var_map sparad som: {var_map_json}")


if __name__ == "__main__":
    main()